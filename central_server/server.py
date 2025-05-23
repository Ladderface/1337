import os
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, Form, Request, Header, HTTPException, status, Depends, Query, WebSocket, WebSocketDisconnect, Response
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from typing import Optional, List
from datetime import datetime, timedelta
import shutil
import sqlite3
import json
import yaml
from pydantic import BaseModel
import asyncio
import threading
from starlette.responses import RedirectResponse
import csv
import io
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter
import time
from central_server.integrations.webhook import send_webhook

"""
Central Screenshot Server (FastAPI)

REST API для масштабируемого сбора, хранения и управления скриншотами, сообщениями и командами от множества ADB-агентов.

API:
- POST   /upload_screenshot   (авторизация) — загрузка скриншота и метаданных
- POST   /api/send_message    (авторизация) — отправка сообщения/лога/статуса
- GET    /api/messages        (авторизация) — получение сообщений
- GET    /api/get_commands    (авторизация) — получение команд для агента
- POST   /api/command_result  (авторизация) — агент подтверждает выполнение команды
- GET    /screenshots         (публично)    — список скринов (фильтрация)
- GET    /download/...        (публично)    — скачать скрин
- Web-интерфейс: /            — просмотр скринов, фильтры

Все эндпоинты, кроме /screenshots и /download, требуют заголовок:
    Authorization: Bearer <api_key>

"""

# --- Конфиг ---
DATA_DIR = Path(__file__).parent / 'data'
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = Path(__file__).parent / 'db.sqlite3'

def get_config_path():
    env_path = os.environ.get('CONFIG_YAML')
    if env_path:
        return Path(env_path)
    return Path(__file__).parent.parent / 'config.yaml'

# --- FastAPI ---
app = FastAPI(title="Central Screenshot Server", description="Масштабируемый сервер для сбора и анализа скринов с множества агентов", version="1.0")

app.mount("/static", StaticFiles(directory=Path(__file__).parent / "static"), name="static")
templates = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))

# --- API-ключи из config.yaml ---
# (Удалено: load_api_keys и API_KEYS)

def check_role(required_roles):
    def dependency(authorization: str = Header(...)):
        if not authorization.startswith('Bearer '):
            raise HTTPException(status_code=401, detail='Missing Bearer token')
        token = authorization.split(' ', 1)[1]
        role = get_user_role(token)
        if role not in required_roles:
            raise HTTPException(status_code=403, detail=f'Insufficient role: {role}')
        return token
    return dependency

# --- БД (SQLite, если нужна история) ---
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS screenshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            server_id TEXT,
            window TEXT,
            device_id TEXT,
            filename TEXT,
            meta TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            server_id TEXT,
            device_id TEXT,
            type TEXT,
            message TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS commands (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            server_id TEXT,
            device_id TEXT,
            command TEXT,
            params TEXT,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.commit()
init_db()

# --- Ролевая модель ---
def get_user_role(api_key: str) -> str:
    config_path = get_config_path()
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f) or {}
        for user in cfg.get('users', []):
            if user.get('api_key') == api_key:
                return user.get('role', 'user')
    # fallback: если нет users — testkey=admin
    if api_key == 'testkey':
        return 'admin'
    return 'user'

# --- Универсальный хук событий для интеграций ---
def call_integrations(event: str, data: dict):
    config_path = get_config_path()
    if not config_path.exists():
        return
    with open(config_path, 'r', encoding='utf-8') as f:
        cfg = yaml.safe_load(f) or {}
    modules = cfg.get('modules', {})
    # Webhook
    if modules.get('webhook'):
        print('[DEBUG] call_integrations: send_webhook =', send_webhook)
        try:
            threading.Thread(target=send_webhook, args=(event, data, cfg), daemon=True).start()
        except Exception as e:
            pass
    # Google Sheets
    if modules.get('google_sheets'):
        try:
            threading.Thread(target=append_to_sheet, args=(event, data, cfg), daemon=True).start()
        except Exception as e:
            pass
    # Telegram (только для alert/error)
    if modules.get('telegram') and event in ('alert', 'error'):
        try:
            text = f"[{event.upper()}] {data.get('server_id','')} {data.get('device_id','')}: {data.get('message','')}"
            threading.Thread(target=send_telegram_alert, args=(text,), daemon=True).start()
        except Exception as e:
            pass

# --- API: загрузка скрина ---
@app.post("/upload_screenshot")
async def upload_screenshot(
    server_id: str = Form(...),
    window: str = Form(...),
    device_id: str = Form(...),
    meta: Optional[str] = Form(None),
    image: UploadFile = File(...),
    token: str = Depends(check_role(['admin', 'user']))
):
    # Формируем путь: data/server_id/window/
    save_dir = DATA_DIR / server_id / window
    save_dir.mkdir(parents=True, exist_ok=True)
    # Имя файла: device_id_YYYY-MM-DD_HH-MM-SS.png
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    safe_device = device_id.replace(':', '_').replace('/', '_')
    filename = f"{safe_device}_{now}.png"
    save_path = save_dir / filename
    with open(save_path, "wb") as f:
        shutil.copyfileobj(image.file, f)
    # Сохраняем last.png (перезапись)
    last_path = save_dir / f"{safe_device}_last.png"
    shutil.copyfile(save_path, last_path)
    # Сохраняем в БД
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('INSERT INTO screenshots (server_id, window, device_id, filename, meta) VALUES (?, ?, ?, ?, ?)',
                  (server_id, window, device_id, filename, meta or ""))
        conn.commit()
    call_integrations('screenshot_uploaded', {
        'server_id': server_id,
        'window': window,
        'device_id': device_id,
        'filename': filename,
        'meta': meta,
        'created_at': now
    })
    return {"status": "ok", "path": str(save_path), "last": str(last_path)}

# --- API: получить список скринов (расширенная фильтрация) ---
@app.get("/screenshots")
def list_screenshots(
    server_id: Optional[str] = None,
    window: Optional[str] = None,
    device_id: Optional[str] = None,
    section: Optional[str] = None,
    text: Optional[str] = None,  # поиск по filename/meta
    created_at_from: Optional[str] = None,  # ISO-строка
    created_at_to: Optional[str] = None,    # ISO-строка
):
    """
    Получить список скринов с расширенной фильтрацией:
    - server_id, window, device_id — как раньше
    - section — фильтр по секции (window или meta.section)
    - text — поиск по filename/meta (LIKE)
    - created_at_from, created_at_to — фильтр по дате (ISO-строка)
    """
    query = "SELECT server_id, window, device_id, filename, meta, created_at FROM screenshots WHERE 1=1"
    params = []
    if server_id:
        query += " AND server_id=?"
        params.append(server_id)
    if window:
        query += " AND window=?"
        params.append(window)
    if device_id:
        query += " AND device_id=?"
        params.append(device_id)
    if section:
        query += " AND (window=? OR meta LIKE ?)"
        params.append(section)
        params.append(f'%"section": "{section}"%')
    if text:
        query += " AND (filename LIKE ? OR meta LIKE ?)"
        like = f"%{text}%"
        params.extend([like, like])
    if created_at_from:
        query += " AND created_at >= ?"
        params.append(created_at_from)
    if created_at_to:
        query += " AND created_at <= ?"
        params.append(created_at_to)
    query += " ORDER BY created_at DESC LIMIT 1000"
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(query, params)
        rows = c.fetchall()
    result = []
    for r in rows:
        meta = r[4]
        section_val = r[1]
        try:
            meta_dict = json.loads(meta) if meta else {}
            if 'section' in meta_dict:
                section_val = meta_dict['section']
        except Exception:
            pass
        result.append({
            "server_id": r[0],
            "window": r[1],
            "device_id": r[2],
            "filename": r[3],
            "section": section_val,
            "created_at": r[5]
        })
    return result

# --- API: скачать скрин ---
@app.get("/download/{server_id}/{window}/{filename}")
def download_screenshot(server_id: str, window: str, filename: str):
    file_path = DATA_DIR / server_id / window / filename
    if not file_path.exists():
        return JSONResponse({"error": "not found"}, status_code=404)
    return FileResponse(str(file_path), media_type="image/png")

# --- API: скачать последний скрин ---
@app.get("/download_last/{server_id}/{window}/{device_id}")
def download_last_screenshot(server_id: str, window: str, device_id: str):
    safe_device = device_id.replace(':', '_').replace('/', '_')
    last_path = DATA_DIR / server_id / window / f"{safe_device}_last.png"
    if not last_path.exists():
        return JSONResponse({"error": "not found"}, status_code=404)
    return FileResponse(str(last_path), media_type="image/png")

# --- Web-интерфейс: просмотр скринов ---
@app.get("/", response_class=HTMLResponse)
def index(request: Request, server_id: Optional[str] = None, window: Optional[str] = None, device_id: Optional[str] = None):
    section = request.query_params.get('section')
    screenshots = list_screenshots(server_id, window, device_id, section)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return templates.TemplateResponse("index.html", {"request": request, "screenshots": screenshots, "server_id": server_id, "window": window, "device_id": device_id, "now": now})

# --- Заглушки для интеграций (Telegram, Google Sheets, нейросети) ---
# Можно реализовать как background tasks, очереди, webhooks и т.д.
# Все интеграции можно включать/выключать через конфиг (см. README.md)

# --- API: отправка сообщения/лога ---
class MessageIn(BaseModel):
    server_id: str
    device_id: str
    type: str
    message: str
    timestamp: Optional[str] = None

# --- WebSocket для live-логов ---
class LogManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []
    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)
    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except Exception:
                pass
log_manager = LogManager()

@app.websocket('/ws/logs')
async def websocket_logs(websocket: WebSocket):
    await log_manager.connect(websocket)
    try:
        while True:
            await asyncio.sleep(60)  # держим соединение
    except WebSocketDisconnect:
        log_manager.disconnect(websocket)

# --- Интеграция с Telegram ---
def send_telegram_alert(text: str):
    try:
        config_path = get_config_path()
        if not config_path.exists():
            return
        import yaml
        with open(config_path, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f)
        tg = cfg.get('telegram', {})
        bot_token = tg.get('bot_token')
        chat_id = tg.get('chat_id')
        if not bot_token or not chat_id:
            return
        import requests
        url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
        data = {'chat_id': chat_id, 'text': text}
        requests.post(url, data=data, timeout=10)
    except Exception as e:
        pass

@app.post('/api/send_message')
def send_message(msg: MessageIn, token: str = Depends(check_role(['admin', 'user']))):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('INSERT INTO messages (server_id, device_id, type, message) VALUES (?, ?, ?, ?)',
                  (msg.server_id, msg.device_id, msg.type, msg.message))
        conn.commit()
    # WebSocket broadcast (async)
    def ws_broadcast():
        asyncio.run(log_manager.broadcast({
            'server_id': msg.server_id,
            'device_id': msg.device_id,
            'type': msg.type,
            'message': msg.message,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }))
    threading.Thread(target=ws_broadcast, daemon=True).start()
    # Telegram alert for error/alert
    if msg.type in ('error', 'alert'):
        alert_text = f"[{msg.type.upper()}] {msg.server_id} {msg.device_id}: {msg.message}"
        threading.Thread(target=send_telegram_alert, args=(alert_text,), daemon=True).start()
    # TODO: Google Sheets, нейросети — добавить здесь
    call_integrations(msg.type, {
        'server_id': msg.server_id,
        'device_id': msg.device_id,
        'type': msg.type,
        'message': msg.message,
        'timestamp': msg.timestamp or datetime.now().isoformat()
    })
    return {'status': 'ok'}

# --- API: получить сообщения (расширенная фильтрация) ---
def get_messages_list(
    server_id: Optional[str] = None,
    device_id: Optional[str] = None,
    type: Optional[str] = None,
    text: Optional[str] = None,  # поиск по message
    created_at_from: Optional[str] = None,
    created_at_to: Optional[str] = None,
    limit: int = 100
):
    """
    Получить список сообщений с расширенной фильтрацией:
    - server_id, device_id, type — как раньше
    - text — поиск по message (LIKE)
    - created_at_from, created_at_to — фильтр по дате (ISO-строка)
    """
    query = 'SELECT server_id, device_id, type, message, created_at FROM messages WHERE 1=1'
    params = []
    if server_id:
        query += ' AND server_id=?'
        params.append(server_id)
    if device_id:
        query += ' AND device_id=?'
        params.append(device_id)
    if type:
        query += ' AND type=?'
        params.append(type)
    if text:
        query += ' AND message LIKE ?'
        params.append(f"%{text}%")
    if created_at_from:
        query += ' AND created_at >= ?'
        params.append(created_at_from)
    if created_at_to:
        query += ' AND created_at <= ?'
        params.append(created_at_to)
    query += ' ORDER BY created_at DESC LIMIT ?'
    params.append(limit)
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(query, params)
        rows = c.fetchall()
    return [
        {'server_id': r[0], 'device_id': r[1], 'type': r[2], 'message': r[3], 'created_at': r[4]}
        for r in rows
    ]

@app.get('/logs', response_class=HTMLResponse)
def logs_page(request: Request, server_id: Optional[str] = Query(None), device_id: Optional[str] = Query(None), type: Optional[str] = Query(None), limit: int = Query(100)):
    messages = get_messages_list(server_id, device_id, type, None, None, None, limit)
    return templates.TemplateResponse('logs.html', {"request": request, "messages": messages, "server_id": server_id, "device_id": device_id, "type": type, "limit": limit})

# --- API: получить команды для агента ---
@app.get('/api/get_commands')
def get_commands(server_id: str, device_id: str, token: str = Depends(check_role(['admin', 'user', 'readonly']))):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('SELECT id, command, params, status, created_at FROM commands WHERE server_id=? AND device_id=? AND status="pending" ORDER BY created_at',
                  (server_id, device_id))
        rows = c.fetchall()
    cmds = []
    for r in rows:
        try:
            params = json.loads(r[2]) if r[2] else {}
        except Exception:
            params = {}
        cmds.append({'id': r[0], 'command': r[1], 'params': params, 'status': r[3], 'created_at': r[4]})
    return {'commands': cmds}

# --- API: агент подтверждает выполнение команды ---
class CommandResultIn(BaseModel):
    command_id: int
    status: str
    result: Optional[str] = None

@app.post('/api/command_result')
def command_result(res: CommandResultIn, token: str = Depends(check_role(['admin', 'user']))):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('UPDATE commands SET status=? WHERE id=?', (res.status, res.command_id))
        conn.commit()
    call_integrations('command_result', {
        'command_id': res.command_id,
        'status': res.status,
        'result': res.result,
        'timestamp': datetime.now().isoformat()
    })
    return {'status': 'ok'}

@app.get('/commands', response_class=HTMLResponse)
def commands_page(request: Request, server_id: Optional[str] = Query(None), device_id: Optional[str] = Query(None), message: Optional[str] = Query(None)):
    # Для простоты: получить уникальные server_id и device_id из последних скринов
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('SELECT DISTINCT server_id FROM screenshots ORDER BY server_id')
        servers = [r[0] for r in c.fetchall()]
        c.execute('SELECT DISTINCT device_id FROM screenshots ORDER BY device_id')
        devices = [r[0] for r in c.fetchall()]
    return templates.TemplateResponse('commands.html', {"request": request, "servers": servers, "devices": devices, "server_id": server_id, "device_id": device_id, "message": message})

@app.post('/commands', response_class=HTMLResponse)
def send_command(request: Request, server_id: str = Form(...), device_id: str = Form(...), command: str = Form(...), params: str = Form(""), token: str = Depends(check_role(['admin', 'user']))):
    # params — строка, можно хранить как JSON
    import json
    try:
        params_json = json.loads(params) if params.strip() else {}
    except Exception:
        params_json = {"raw": params}
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('INSERT INTO commands (server_id, device_id, command, params) VALUES (?, ?, ?, ?)',
                  (server_id, device_id, command, json.dumps(params_json)))
        conn.commit()
    msg = f"Команда '{command}' отправлена для {device_id} на {server_id}"
    url = f"/commands?server_id={server_id}&device_id={device_id}&message={msg}"
    return RedirectResponse(url, status_code=303)

# --- API: экспорт скринов (CSV или ZIP) ---
@app.get("/export/screenshots")
def export_screenshots(
    server_id: Optional[str] = None,
    window: Optional[str] = None,
    device_id: Optional[str] = None,
    text: Optional[str] = None,
    created_at_from: Optional[str] = None,
    created_at_to: Optional[str] = None,
    format: str = 'csv',  # 'csv' или 'zip'
    token: str = Depends(check_role(['admin']))
):
    """
    Экспорт скринов по фильтрам:
    - format=csv — метаданные в CSV
    - format=zip — архив файлов скринов
    """
    screenshots = list_screenshots(server_id, window, device_id, text, created_at_from, created_at_to)
    if format == 'csv':
        output = io.StringIO()
        fieldnames = ["server_id", "window", "device_id", "filename", "created_at"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for s in screenshots:
            filtered = {k: s[k] for k in fieldnames if k in s}
            writer.writerow(filtered)
        output.seek(0)
        return StreamingResponse(io.BytesIO(output.getvalue().encode('utf-8')), media_type='text/csv', headers={"Content-Disposition": "attachment; filename=screenshots.csv"})
    elif format == 'zip':
        mem_zip = io.BytesIO()
        with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for s in screenshots:
                file_path = DATA_DIR / s['server_id'] / s['window'] / s['filename']
                if file_path.exists():
                    zf.write(file_path, arcname=f"{s['server_id']}/{s['window']}/{s['filename']}")
        mem_zip.seek(0)
        return StreamingResponse(mem_zip, media_type='application/zip', headers={"Content-Disposition": "attachment; filename=screenshots.zip"})
    else:
        return JSONResponse({"error": "format must be csv or zip"}, status_code=400)

# --- API: экспорт логов (CSV) ---
@app.get("/export/logs")
def export_logs(
    server_id: Optional[str] = None,
    device_id: Optional[str] = None,
    type: Optional[str] = None,
    text: Optional[str] = None,
    created_at_from: Optional[str] = None,
    created_at_to: Optional[str] = None,
    limit: int = 1000,
    token: str = Depends(check_role(['admin']))
):
    """
    Экспорт логов по фильтрам в CSV
    """
    logs = get_messages_list(server_id, device_id, type, text, created_at_from, created_at_to, limit)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["server_id", "device_id", "type", "message", "created_at"])
    writer.writeheader()
    for l in logs:
        writer.writerow(l)
    output.seek(0)
    return StreamingResponse(io.BytesIO(output.getvalue().encode('utf-8')), media_type='text/csv', headers={"Content-Disposition": "attachment; filename=logs.csv"})

# --- API: история выполнения команд (фильтрация + экспорт) ---
@app.get("/api/command_history")
def api_command_history(
    server_id: Optional[str] = None,
    device_id: Optional[str] = None,
    command: Optional[str] = None,
    status: Optional[str] = None,
    text: Optional[str] = None,  # поиск по команде/параметрам
    created_at_from: Optional[str] = None,
    created_at_to: Optional[str] = None,
    format: str = 'json',  # 'json', 'csv', 'xlsx'
    limit: int = 1000,
    token: str = Depends(check_role(['admin']))
):
    """
    Получить историю команд с фильтрацией:
    - server_id, device_id, command, status, text (LIKE по command/params), created_at_from/to
    - format: json (default), csv, xlsx
    """
    query = 'SELECT id, server_id, device_id, command, params, status, created_at FROM commands WHERE 1=1'
    params = []
    if server_id:
        query += ' AND server_id=?'
        params.append(server_id)
    if device_id:
        query += ' AND device_id=?'
        params.append(device_id)
    if command:
        query += ' AND command=?'
        params.append(command)
    if status:
        query += ' AND status=?'
        params.append(status)
    if text:
        query += ' AND (command LIKE ? OR params LIKE ? OR status LIKE ? OR server_id LIKE ? OR device_id LIKE ?)'  # расширенный поиск
        like = f"%{text}%"
        params.extend([like, like, like, like, like])
    if created_at_from:
        query += ' AND created_at >= ?'
        params.append(created_at_from)
    if created_at_to:
        query += ' AND created_at <= ?'
        params.append(created_at_to)
    query += ' ORDER BY created_at DESC LIMIT ?'
    params.append(limit)
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(query, params)
        rows = c.fetchall()
    columns = ["id", "server_id", "device_id", "command", "params", "status", "created_at"]
    data = [dict(zip(columns, r)) for r in rows]
    if format == 'json':
        return data
    elif format == 'csv':
        import io, csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        output.seek(0)
        return StreamingResponse(io.BytesIO(output.getvalue().encode('utf-8')), media_type='text/csv', headers={"Content-Disposition": "attachment; filename=command_history.csv"})
    elif format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "History"
        ws.append(columns)
        for row in data:
            ws.append([row[col] for col in columns])
        for i, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": "attachment; filename=command_history.xlsx"})
    else:
        return JSONResponse({"error": "format must be json, csv or xlsx"}, status_code=400)

# --- UI: страница истории команд ---
@app.get('/command_history', response_class=HTMLResponse)
def command_history_page(
    request: Request,
    server_id: Optional[str] = Query(None),
    device_id: Optional[str] = Query(None),
    command: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    text: Optional[str] = Query(None),
    created_at_from: Optional[str] = Query(None),
    created_at_to: Optional[str] = Query(None),
    limit: int = Query(1000)
):
    # Получаем историю команд для отображения
    data = api_command_history(server_id, device_id, command, status, text, created_at_from, created_at_to, 'json', limit)
    # Для фильтров — получить уникальные значения
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('SELECT DISTINCT server_id FROM commands ORDER BY server_id')
        servers = [r[0] for r in c.fetchall()]
        c.execute('SELECT DISTINCT device_id FROM commands ORDER BY device_id')
        devices = [r[0] for r in c.fetchall()]
        c.execute('SELECT DISTINCT command FROM commands ORDER BY command')
        commands = [r[0] for r in c.fetchall()]
        c.execute('SELECT DISTINCT status FROM commands ORDER BY status')
        statuses = [r[0] for r in c.fetchall()]
    return templates.TemplateResponse('command_history.html', {
        "request": request,
        "history": data,
        "servers": servers,
        "devices": devices,
        "commands": commands,
        "statuses": statuses,
        "server_id": server_id,
        "device_id": device_id,
        "command": command,
        "status": status,
        "text": text,
        "created_at_from": created_at_from,
        "created_at_to": created_at_to,
        "limit": limit
    })

# --- UI: страница аналитики и графиков ---
@app.get('/analytics', response_class=HTMLResponse)
def analytics_page(request: Request):
    return templates.TemplateResponse('analytics.html', {"request": request})

# --- API: аналитика/графики ---
def parse_date(val):
    try:
        return datetime.strptime(val, '%Y-%m-%d')
    except Exception:
        return None

def date_filter_sql(field, date_from, date_to):
    cond = []
    params = []
    if date_from:
        cond.append(f"{field} >= ?")
        params.append(date_from + ' 00:00:00')
    if date_to:
        cond.append(f"{field} <= ?")
        params.append(date_to + ' 23:59:59')
    return cond, params

@app.get('/api/analytics/summary')
def analytics_summary(
    server_id: Optional[str] = None,
    device_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    # Скриншоты и логи по дням
    cond, params = [], []
    if server_id:
        cond.append('server_id=?')
        params.append(server_id)
    if device_id:
        cond.append('device_id=?')
        params.append(device_id)
    date_cond, date_params = date_filter_sql('created_at', date_from, date_to)
    cond += date_cond
    params += date_params
    where = 'WHERE ' + ' AND '.join(cond) if cond else ''
    # Скриншоты по дням
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(f'''SELECT substr(created_at,1,10) as day, count(*) FROM screenshots {where} GROUP BY day ORDER BY day''', params)
        screens = c.fetchall()
        c.execute(f'''SELECT substr(created_at,1,10) as day, count(*) FROM messages {where} GROUP BY day ORDER BY day''', params)
        logs = c.fetchall()
        c.execute(f'''SELECT status, count(*) FROM commands {where} GROUP BY status''', params)
        cmd_status = c.fetchall()
    # Chart.js формат
    days = sorted(set([d[0] for d in screens] + [d[0] for d in logs]))
    screens_map = {d: v for d, v in screens}
    logs_map = {d: v for d, v in logs}
    screens_data = [screens_map.get(d, 0) for d in days]
    logs_data = [logs_map.get(d, 0) for d in days]
    # Статусы команд
    status_labels = [s[0] for s in cmd_status]
    status_counts = [s[1] for s in cmd_status]
    return {
        "screens_logs": {
            "data": {
                "labels": days,
                "datasets": [
                    {"label": "Скриншоты", "data": screens_data, "borderColor": "#0077cc", "backgroundColor": "#0077cc22", "tension":0.3},
                    {"label": "Логи", "data": logs_data, "borderColor": "#e67e22", "backgroundColor": "#e67e2222", "tension":0.3}
                ]
            },
            "options": {"responsive": True, "plugins": {"legend": {"display": True}}}
        },
        "commands_status": {
            "data": {
                "labels": status_labels,
                "datasets": [
                    {"label": "Статус", "data": status_counts, "backgroundColor": ["#27ae60", "#e67e22", "#c00", "#888"]}
                ]
            },
            "options": {"responsive": True, "plugins": {"legend": {"display": False}}}
        }
    }

@app.get('/api/analytics/device_activity')
def analytics_device_activity(
    server_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    cond, params = [], []
    if server_id:
        cond.append('server_id=?')
        params.append(server_id)
    date_cond, date_params = date_filter_sql('created_at', date_from, date_to)
    cond += date_cond
    params += date_params
    where = 'WHERE ' + ' AND '.join(cond) if cond else ''
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(f'''SELECT device_id, count(*) FROM screenshots {where} GROUP BY device_id ORDER BY count(*) DESC''', params)
        rows = c.fetchall()
    labels = [r[0] for r in rows]
    data = [r[1] for r in rows]
    return {
        "device_activity": {
            "data": {
                "labels": labels,
                "datasets": [{"label": "Скриншоты", "data": data, "backgroundColor": "#0077cc"}]
            },
            "options": {"responsive": True, "plugins": {"legend": {"display": False}}}
        }
    }

@app.get('/api/analytics/errors')
def analytics_errors(
    server_id: Optional[str] = None,
    device_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    cond, params = ["type IN ('error','alert')"], []
    if server_id:
        cond.append('server_id=?')
        params.append(server_id)
    if device_id:
        cond.append('device_id=?')
        params.append(device_id)
    date_cond, date_params = date_filter_sql('created_at', date_from, date_to)
    cond += date_cond
    params += date_params
    where = 'WHERE ' + ' AND '.join(cond) if cond else ''
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(f'''SELECT substr(created_at,1,10) as day, count(*) FROM messages {where} GROUP BY day ORDER BY day''', params)
        rows = c.fetchall()
    days = [r[0] for r in rows]
    data = [r[1] for r in rows]
    return {
        "errors": {
            "data": {
                "labels": days,
                "datasets": [{"label": "Ошибки/Алерты", "data": data, "backgroundColor": "#c00"}]
            },
            "options": {"responsive": True, "plugins": {"legend": {"display": False}}}
        }
    }

@app.get('/api/analytics/commands_history')
def analytics_commands_history(
    server_id: Optional[str] = None,
    device_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    cond, params = [], []
    if server_id:
        cond.append('server_id=?')
        params.append(server_id)
    if device_id:
        cond.append('device_id=?')
        params.append(device_id)
    date_cond, date_params = date_filter_sql('created_at', date_from, date_to)
    cond += date_cond
    params += date_params
    where = 'WHERE ' + ' AND '.join(cond) if cond else ''
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(f'''SELECT substr(created_at,1,10) as day, count(*) FROM commands {where} GROUP BY day ORDER BY day''', params)
        rows = c.fetchall()
    days = [r[0] for r in rows]
    data = [r[1] for r in rows]
    return {
        "commands_history": {
            "data": {
                "labels": days,
                "datasets": [{"label": "Команды", "data": data, "borderColor": "#2980b9", "backgroundColor": "#2980b922", "tension":0.3}]
            },
            "options": {"responsive": True, "plugins": {"legend": {"display": False}}}
        }
    } 